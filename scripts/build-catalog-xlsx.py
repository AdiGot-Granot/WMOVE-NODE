import csv, os, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV = os.path.join(BASE, "catalog", "endpoint-catalog.csv")
OUT = os.path.join(BASE, "catalog", "wmove-endpoint-catalog.xlsx")

rows = list(csv.DictReader(open(CSV, encoding="utf-8")))

def is_variant(f):
    return bool(re.search(r"(_FIX|-YOSSI|\(\d+\))", f, re.I)) or f.upper() in (
        "MPPROC-YOSSI.PRG", "MPPROC_FIX.PRG", "MPESTX.PRG")

for r in rows:
    r["variant"] = "yes" if is_variant(r["sourceFile"]) else ""

FONT = "Arial"
hdr_fill = PatternFill("solid", fgColor="1F3864")
hdr_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
title_font = Font(name=FONT, bold=True, size=14, color="1F3864")
sub_font = Font(name=FONT, italic=True, size=9, color="595959")
cell_font = Font(name=FONT, size=10)
var_fill = PatternFill("solid", fgColor="FCE4D6")
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ── Backlog sheet ──────────────────────────────────────────────────────
ws = wb.active
ws.title = "Backlog"
cols = ["module","domain","method","kind","sourceFile","line","variant","status","priority","notes"]
heads = ["Module","Domain","Method","Kind","Source File","Line","Variant/Dupe","Status","Priority","Notes"]
ws.append(heads)
for c in range(1, len(heads)+1):
    cell = ws.cell(1, c); cell.fill = hdr_fill; cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = border
for r in rows:
    ws.append([r["module"], r["domain"], r["method"], r["kind"], r["sourceFile"],
               int(r["line"]), r["variant"], r["status"], r["priority"], r["notes"]])
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(cols)):
    for cell in row:
        cell.font = cell_font; cell.border = border
    if row[6].value == "yes":
        for cell in row: cell.fill = var_fill
widths = [13,22,26,11,26,7,13,10,10,30]
for i,w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"

# ── Summary sheet ──────────────────────────────────────────────────────
s = wb.create_sheet("Summary", 0)
s["A1"] = "wMove — Endpoint Migration Backlog"; s["A1"].font = title_font
s["A2"] = "Auto-generated from legacy VFP process classes (MP*/MA*). Counts via COUNTIF on the Backlog sheet."
s["A2"].font = sub_font
s.append([])

order = [
    ("MPLOGIN","Auth / login",1,"Foundation — needed by everything"),
    ("MPDIR","Directory / distance",2,"Low risk, proves the framework"),
    ("MPUTIL","Utilities / config",2,"Mostly read-only reference"),
    ("MPLEADS","Leads",3,"High value, self-contained"),
    ("MPAGENT","Agents",4,""),
    ("MPEST","Estimates",5,"POC done (inventory page)"),
    ("MPINV","Inventory",5,""),
    ("MPOPR","Operations / jobs",6,"Large"),
    ("MPSTG","Storage",7,""),
    ("MPCHARGE","Billing / card & e-check",8,"HIGHEST RISK — migrate late"),
    ("MPREP","Reports",9,""),
    ("MPMAINTEN","Maintenance",9,""),
    ("MPAUTO","Auto transport",9,""),
    ("MPADMIN","Admin",9,""),
    ("MPXML","XML / API",9,"With batch jobs"),
]
hr = 4
heads2 = ["Phase","Module","Domain","Endpoints (live)","Notes"]
s.append(heads2)
for c in range(1,6):
    cell = s.cell(hr, c); cell.fill = hdr_fill; cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center"); cell.border = border
for mod, dom, ph, note in order:
    s.append([ph, mod, dom, f'=COUNTIFS(Backlog!A:A,B{s.max_row+1},Backlog!G:G,"")', note])
for row in s.iter_rows(min_row=hr+1, max_row=s.max_row, max_col=5):
    for cell in row: cell.font = cell_font; cell.border = border
tot = s.max_row + 1
s.cell(tot,3,"TOTAL live endpoints").font = Font(name=FONT, bold=True)
s.cell(tot,4,'=COUNTIF(Backlog!G2:G100000,"")').font = Font(name=FONT, bold=True)
s.cell(tot+1,3,"Variant/duplicate (triage out)").font = Font(name=FONT, italic=True)
s.cell(tot+1,4,'=COUNTIF(Backlog!G2:G100000,"yes")').font = Font(name=FONT, italic=True)
for i,w in enumerate([8,14,24,16,34],1): s.column_dimensions[get_column_letter(i)].width = w

wb.save(OUT)
print("wrote", OUT, "rows:", len(rows))
